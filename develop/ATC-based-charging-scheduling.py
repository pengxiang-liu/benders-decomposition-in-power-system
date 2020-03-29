# -----------------------------------------------------------------------------
# Copyright 2020, Southeast University, Liu Pengxiang
# 
# A EV charging scheduling model using constraint programming technology.
# The model will be upgrated as a hybrid mathematical and constraint 
# programming model.
# 
# EV data can be downloaded from: https://pod-point.com/guides/vehicles
# -----------------------------------------------------------------------------


import sys
import math
import numpy as np
import gurobipy as gp
import openpyxl as pyxl
import docplex.cp.model as cp
import matplotlib.pyplot as plt
import docplex.cp.utils_visu as visu


# This class creates the parameter class
# 
class Parameter(object):

    # Initialization
    def __init__(self, filename, flag, number):
        '''Pre-processing'''
        '''-----------------------------------------------------------------'''
        Excl = Excel()  # excel file execution tool
        Data = Excl.read_file(filename)
        self.N_mins = 5     # number of minutes for each time section
        self.N_sect = int(24 * 60 / self.N_mins)  # number of time section
        self.C_rate = [50, 22, 7]  # charging rate
        self.N_rate = len(self.C_rate)  # number of charging rate types
        if flag == 1:
            self.charging_demand(Data, number, Excl, filename)
            Data = Excl.read_file(filename)  # re-read
        '''Data input'''
        '''-----------------------------------------------------------------'''
        # 0) Line
        self.Line   = Data[0]  # line
        self.N_line = len(self.Line)  # number of line
        # 1) Bus
        self.Bus    = Data[1]  # bus
        self.N_bus  = len(self.Bus )  # number of bus
        # 2) Sub
        self.Sub    = Data[2]  # substation
        self.N_sub  = len(self.Sub )  # number of substation
        # 3) Gen
        self.Gen    = Data[3]  # renewable generation
        self.N_gen  = len(self.Gen )  # number of renewables
        # 4) Day
        self.Hour   = Data[4]  # data of load, wind and solar station
        self.N_hour = len(self.Hour)  # number of hours
        # 5) Time
        self.Time   = Data[5]  # arrival and departure distribution
        self.N_time = len(self.Time)  # number of charging section
        # 6) Electric vehicle
        self.Batr   = Data[6]  # information of EV battery
        self.N_Batr = len(self.Batr)  # number of battery types
        # 7) Charging Park
        self.Park   = Data[7]  # charging park
        self.N_park = len(self.Park)  # number of charging park
        # 8) Charging pile
        self.Pile   = Data[8]  # information of charging pile
        self.N_pile = len(self.Pile)  # number of charging pile
        # 9) Charging demand
        self.Ecar   = Data[9]  # charging demand of EV
        self.N_ecar = len(self.Ecar)  # number of EV
        '''Modeling parameters'''
        '''-----------------------------------------------------------------'''
        # 1) System
        self.Base_V = 12.66  # voltage: 12.66 kV
        self.Base_S = 10.00  # power:   10.00 MVA
        self.Base_Z = self.Base_V ** 2 / self.Base_S  # impedance
        self.Base_I = self.Base_S / self.Base_V / np.sqrt(3)  # current
        self.Factor = 0.31756  # power factor (rad)
        self.N_scen = 4  # number of reconfiguration scenarios
        # 2) Cost and price
        self.Cost_pen = 500  # cost of load shedding
        self.Cost_los = 25   # cost of energy loss
        self.Cost_alt = 50   # cost of charging place alteration
        self.Cost_out = 100  # cost of timeout per N_mins
        # 3) Other
        self.Big_M = 1e2  # a sufficient large number
        self.V_min = (0.95 * self.Base_V) ** 2
        self.V_max = (1.05 * self.Base_V) ** 2
        # 4) Bus-Line connectivity
        self.Line_head = [[] for i in range(self.N_bus)]
        self.Line_tail = [[] for i in range(self.N_bus)]
        for i in range(self.N_line):
            head = self.Line[i][1]
            tail = self.Line[i][2]
            self.Line_head[int(round(head))].append(i)
            self.Line_tail[int(round(tail))].append(i)
    
    # charging demand
    def charging_demand(self, Data, number, excel, filename):
        info_time = Data[5]  # time
        info_batr = Data[6]  # battery
        demand = np.zeros((number, 9))
        for i in range(number):
            # Determine charging time
            while True:
                # time of arrival and departure
                hour_arr = self.roulette(info_time[:, 1])
                hour_dep = self.roulette(info_time[:, 2])
                # offset
                extr_arr = np.random.randint(0, self.N_sect / 24 + 0.1)  
                extr_dep = np.random.randint(0, self.N_sect / 24 + 0.1)
                # actual
                time_arr = int(hour_arr * self.N_sect / 24 + extr_arr)
                time_dep = int(hour_dep * self.N_sect / 24 + extr_dep)
                # charging at least 20 minutes
                if time_arr <= time_dep - 20 / self.N_mins:
                    break
            # Determine charging demand
            while True:
                # Capacity
                type_num = int(self.roulette(info_batr[:, 6]))
                list_cap = np.zeros(len(self.C_rate))  # charging capacity
                for index, item in enumerate(self.C_rate):
                    time = time_dep - time_arr
                    flag = info_batr[type_num, 2 + index]
                    list_cap[index] = time * flag * item / (self.N_sect / 24)
                capacity = np.max(list_cap)
                # Demand
                average  = info_batr[type_num, 1] / 2
                variance = info_batr[type_num, 1] / 2 / 1.96
                charging = np.random.normal(loc = average, scale = variance)
                # Charging time
                list_tim = np.zeros(len(self.C_rate))  # charging capacity
                for index, rate in enumerate(self.C_rate):
                    flag = info_batr[type_num, 2 + index]
                    time = charging / rate * 60 / self.N_mins
                    time = min(time, time_dep - time_arr)
                    if flag == 1:
                        list_tim[index] = max(int(time), 1)
                    else:
                        list_tim[index] = int(1000)
                if charging > 0 and charging < capacity:
                    break
            demand[i, 0] = int(i)
            demand[i, 1] = time_arr
            demand[i, 2] = time_dep
            demand[i, 3] = type_num
            demand[i, 4] = round(charging, 3)
            demand[i, 5] = np.random.randint(0, 2)
            for index, item in enumerate(list_tim):
                demand[i, 6 + index] = item
        sheetname = 'Ecar'
        title = ["No.", "Arrive", "Leave", "Type", "Demand", "Park", 
                 "Time-50", "Time-22", "Time-07"]
        excel.save_data(filename, sheetname, title, demand)

    # roulette algorithm
    def roulette(self, pdf):
        pdf  = np.append(pdf, 0)
        pick = np.random.rand()
        _sum = 0
        for [index, value] in enumerate(pdf):
            if pick >= _sum:
                _sum = _sum + value
            else:
                break
        solution = index - 1
        return solution


# This class restores the results of MIP power flow model
#
class Result_MIP(object):
    
    # Initialization
    def __init__(self, model, Para, y_line, f_flow, v_flow, l_func, h_cost):
        self.y_line = self.value(y_line, 'int')
        self.f_flow = self.value(f_flow, 'float')
        self.v_flow = self.value(v_flow, 'float')
        self.l_func = self.value(l_func, 'float')
        self.h_cost = self.value(h_cost, 'float')
        self.V_bus  = self.v_flow[N_V_bus  : N_V_bus  + Para.N_bus , :]
        self.P_line = self.v_flow[N_P_line : N_P_line + Para.N_line, :]
        self.Q_line = self.v_flow[N_Q_line : N_Q_line + Para.N_line, :]
        self.P_sub  = self.v_flow[N_P_sub  : N_P_sub  + Para.N_sub , :]
        self.Q_sub  = self.v_flow[N_Q_sub  : N_Q_sub  + Para.N_sub , :]
        self.S_gen  = self.v_flow[N_S_gen  : N_S_gen  + Para.N_gen , :]
        self.C_gen  = self.v_flow[N_C_gen  : N_C_gen  + Para.N_gen , :]
        self.P_cut  = self.v_flow[N_P_cut  : N_P_cut  + Para.N_bus , :]
        self.Q_cut  = self.v_flow[N_Q_cut  : N_Q_cut  + Para.N_bus , :]
        self.P_park = self.v_flow[N_P_park : N_P_park + Para.N_park, :]
    
    # Obtain the value of variables
    def value(self,variable,string):
        # Get value
        key = variable.keys()
        val = variable.copy()
        for i in range(len(key)):
            val[key[i]] = variable[key[i]].x
        # Calculate dimention
        if isinstance(max(key),tuple):  # multi dimention
            dim = tuple([item + 1 for item in max(key)])
        if isinstance(max(key),int):    # one   dimention
            dim = tuple([int(len(key)),1])
        # Convert dictionary to numpy array
        arr = np.zeros(dim, dtype = string)
        for i in range(len(val)):
            arr[key[i]] = val[key[i]]
        return arr


# This class restores the results of COP scheduling model
#
class Result_COP(object):
    
    # Initialization
    def __init__(self, msol, Para, x_ecar, y_ecar, z_ecar, p_park):
        # 1. decision variables
        self.x_ecar = self.value(msol, x_ecar, "itv")
        self.y_ecar = self.value(msol, y_ecar, "itv")
        self.z_ecar = self.value(msol, z_ecar, "itv")
        self.p_park = self.value(msol, p_park, "int" )
        # 2. charging scheduling
        scheduling = np.zeros((Para.N_ecar, Para.N_sect + 3))
        for n in range(Para.N_ecar):
            scheduling[n, 0] = n
            for i in range(Para.N_park):
                if self.y_ecar[n, i].presence == True:
                    scheduling[n, 1] = i
                for k in range(Para.N_rate):
                    if self.z_ecar[n, i, k].presence == True:
                        scheduling[n, 2] = k
            start = self.x_ecar[n].start
            size  = self.x_ecar[n].size
            for i in range(size):
                scheduling[n, start + i + 3] = 1
        self.scheduling = scheduling
        # 3. pile charging and occupation
        charge = np.zeros((Para.N_sect, Para.N_park))
        occupy = np.zeros((Para.N_sect, Para.N_park, Para.N_rate))
        for t in range(Para.N_sect):
            for n in range(Para.N_ecar):
                i = int(scheduling[n, 1])  # park
                k = int(scheduling[n, 2])  # rate
                u = scheduling[n, t + 3]   # occupation
                charge[t, i] = charge[t, i] + Para.C_rate[k] * u
                occupy[t, i, k] = occupy[t, i, k] + u
        self.charge = charge
        self.occupy = occupy
                
    # Obtain the value of variables
    def value(self, msol, var, string):
        # Get value
        sol = {}
        for n in list(var.keys()):
            if string == "int":
                sol[n] = msol.get_var_solution(var[n]).value
            else:
                sol[n] = msol.get_var_solution(var[n])
        # Convert dictionary to numpy array
        return sol
    
    # Save the scheduling results to file
    def save_scheduling(self, schedule):
        excel = Excel()
        filename  = "result/main_scheduling.xlsx"
        sheetname = "charging"
        title = ["No.", "Park", "Rate"]
        for i in range(0, Para.N_hour):
            for j in range(0, 60, Para.N_mins):
                title.append("%02d:%02d" % (i,j))
        excel.save_data(filename, sheetname, title, schedule)


# This class creates the drawing class
#
class Figure(object):

    # Initialization
    def __init__(self, Para):
        # Switch line coordinate
        self.switch_line = []
        self.switch_line.append(np.array([[ 7, 0], [ 7,-2], [3 ,-2], [ 3,-3]]))
        self.switch_line.append(np.array([[ 8, 0], [ 8, 2], [14, 2], [14, 0]]))
        self.switch_line.append(np.array([[11, 0], [11,-3], [ 4,-3]]))
        self.switch_line.append(np.array([[17, 0], [17, 3], [12, 3]]))
        self.switch_line.append(np.array([[ 8, 6], [ 8, 5], [ 8, 3]]))
        # Line coordinate
        self.coordinate = []
        for n in range(Para.N_line):
            if Para.Line[n,6] == 0:  # line
                bus_head = int(round(Para.Line[n, 1]))
                bus_tail = int(round(Para.Line[n, 2]))
                x0 = Para.Bus[bus_head, 3]
                y0 = Para.Bus[bus_head, 4]
                x1 = Para.Bus[bus_tail, 3]
                y1 = Para.Bus[bus_tail, 4]
                self.coordinate.append(np.array([[x0,y0], [x1, y1]]))
            else:
                switch_no = int(Para.Line[n,6] - 1)
                self.coordinate.append(self.switch_line[switch_no])
    
    # Figuring
    def figuring(self, Para, sol, hour):
        self.plot_Bus (Para)
        self.plot_Line(Para, sol, hour)
        plt.axis('equal')
        plt.show()

    # Bus
    def plot_Bus(self, Para):
        for n in range(Para.N_bus):
            plt.plot(Para.Bus[n, 3], Para.Bus[n, 4], 'b.')
            plt.text(Para.Bus[n, 3] + 0.05, Para.Bus[n, 4] + 0.10, '%s' % n)
    
    # Line
    def plot_Line(self, Para, sol, hour):
        scen = int(hour / (24 / Para.N_scen))
        for n in range(Para.N_line):
            for m in range(np.size(self.coordinate[n], 0) - 1):
                x0 = self.coordinate[n][m, 0]
                y0 = self.coordinate[n][m, 1]
                x1 = self.coordinate[n][m + 1, 0]
                y1 = self.coordinate[n][m + 1, 1]
                if sol.y_line[n, scen] == 1:
                    plt.plot([x0, x1], [y0, y1], 'b-' )
                if sol.y_line[n, scen] == 0:
                    plt.plot([x0, x1], [y0, y1], 'r--')


# This class creates the execution tool for Excel files
# 
class Excel(object):

    # Initialization
    def __init__(self):
        pass

    # This function inputs data from Excel files
    def read_file(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i, name in enumerate(book.sheetnames):  # sheet number
            if i < len(book.sheetnames):
                sheet = book[name]
                n_row = sheet.max_row     # number of rows
                n_col = sheet.max_column  # number of columns
                data.append(self.para_filter(sheet, n_row, n_col))
        return data
    
    # This function filters data from the parameters
    def para_filter(self, sheet, n_row, n_col):
        k = 0
        data = []
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if it is a number
                data.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position
                    val = sheet[pos].value
                    if sheet[pos].data_type == 'n':
                        data[k].append(val)
                k = k + 1
        return np.array(data)
    
    # Saving data
    def save_data(self, filename, sheetname, title, data):
        book  = pyxl.load_workbook(filename)
        name  = book.sheetnames[-1]
        book.remove(book[name])
        sheet = book.create_sheet(sheetname)
        sheet.append(title)
        for i in range(len(data)):
            sheet.append(data[i,:].tolist())  # write data
        book.save(filename = filename)


# This class creates the tool for load curtailment detection
# 
class Detect(object):

    # Initialization
    def __init__(self, Para, sol):
        self.flow = -1  # no overload
        self.sub  = -1  # no overload
        P_list = np.where(sol.P_cut > 0)[1]
        Q_list = np.where(sol.Q_cut > 0)[1]
        H_list = np.unique(np.append(P_list, Q_list))
        for i, hour in enumerate(H_list):
            scen = int(hour / (24 / Para.N_scen))
            y_line = sol.y_line[:, scen]
            P_line = sol.P_line[:, hour]
            Q_line = sol.Q_line[:, hour]
            P_sub  = sol.P_sub [:, hour]
            Q_sub  = sol.Q_sub [:, hour]
            # detection
            flow = self.detect_flow(Para, y_line, P_line, Q_line, hour)
            sub  = self.detect_subs(Para, P_sub, Q_sub, hour)
            if i == 0:
                self.flow = flow
                self.sub  = sub
            else:
                if np.size(flow) > 0:
                    self.flow = np.append(self.flow, flow, axis = 0)
                if np.size(sub ) > 0:
                    self.sub  = np.append(self.sub , sub , axis = 0)

    # line flow
    def detect_flow(self, Para, y_line, P_line, Q_line, hour):
        record = []
        for n in range(Para.N_line):
            smax = y_line[n] * Para.Line[n, 3]
            fg_0 = np.abs(P_line[n]) <= smax
            fg_1 = np.abs(Q_line[n]) <= smax
            fg_2 = np.abs(P_line[n] + Q_line[n]) <= 1.414 * smax
            fg_3 = np.abs(P_line[n] - Q_line[n]) <= 1.414 * smax
            if bool(fg_0 and fg_1 and fg_2 and fg_3) == False:
                record.append([hour, Para.Line[n, 1], Para.Line[n, 2]])
        return np.array(record)
    
    # substation
    def detect_subs(self, Para, P_sub, Q_sub, hour):
        record = []
        for n in range(Para.N_sub):
            smax = Para.Sub[n, 2]
            fg_0 = np.abs(P_sub[n]) <= smax
            fg_1 = np.abs(Q_sub[n]) <= smax
            fg_2 = np.abs(P_sub[n] + Q_sub[n]) <= 1.414 * smax
            fg_3 = np.abs(P_sub[n] - Q_sub[n]) <= 1.414 * smax
            if bool(fg_0 and fg_1 and fg_2 and fg_3) == False:
                record.append([hour, Para.Sub[n, 1]])
        return np.array(record)


# This function creates global index
#
def Var_index(Para):

    '''---------------------------Fictitious power flow----------------------------'''
    # 1. Fictitious power flow
    global N_F_line, N_F_sub , N_F_load, N_F_gen , N_F_var
    # Initialization
    N_F_line = 0                       # flow of line
    N_F_load = N_F_line + Para.N_line  # flow of load demand
    N_F_sub  = N_F_load + Para.N_bus   # flow of substation
    N_F_gen  = N_F_sub  + Para.N_sub   # flow of DG
    N_F_var  = N_F_gen  + Para.N_gen   # Number of all fictitious vaariables

    '''------------------------------Real power flow-------------------------------'''
    # 2. Real power flow
    global N_V_bus, N_P_line, N_Q_line, N_P_sub, N_Q_sub
    global N_S_gen, N_C_gen , N_P_cut , N_Q_cut, N_P_park, N_N_var
    # Initialization
    N_V_bus  = 0                       # square of voltage amplitude
    N_P_line = N_V_bus  + Para.N_bus   # power flow (active)
    N_Q_line = N_P_line + Para.N_line  # power flow (reactive)
    N_P_sub  = N_Q_line + Para.N_line  # power injection at substation
    N_Q_sub  = N_P_sub  + Para.N_sub   # power injection at substation
    N_S_gen  = N_Q_sub  + Para.N_sub   # renewables generation
    N_C_gen  = N_S_gen  + Para.N_gen   # renewables curtailment
    N_P_cut  = N_C_gen  + Para.N_gen   # Load shedding (active)
    N_Q_cut  = N_P_cut  + Para.N_bus   # Load shedding (reactive)
    N_P_park = N_Q_cut  + Para.N_bus   # Charging power
    N_N_var  = N_P_park + Para.N_pile  # Number of all variables

    # 3. Augument Lagrangian function
    global N_L_line, N_L_quad, N_L_var
    # Initialization
    N_L_line = 0                       # linear part of ALF
    N_L_quad = N_L_line + Para.N_park  # quadratic part of ALF
    N_L_var  = N_L_quad + Para.N_park  # number of all variables
    
    # Return
    return 0


# This function defines the main function of Analytical Target Cascading (ATC)
# algorithm
# 
def ATC_algorithm(Para):

    # Initial penalty weights
    v = np.zeros((Para.N_park, 2))  # linear terms
    w = np.ones ((Para.N_park, 2))  # quadratic terms
    b = 1.3     # amplification factor

    # Initial target
    sol = COP_scheduling(Para, v[:, 1], w[:, 1], [], 1)

    # Iteration
    target_MIP  = np.zeros((Para.N_park, Para.N_hour))
    for n in range(Para.N_park):
        for h in range(Para.N_hour):
            target_MIP[n, h] = Para.Park[n, 2]
    sol_MIP = MIP_power_flow(Para, v[:, 0], w[:, 0], target_MIP)
    target_COP = sol_MIP.P_park
    sol_COP = COP_scheduling(Para, v[:, 1], w[:, 1], target_COP)
    return 0

# This function defines the MIP_power_flow
#
def MIP_power_flow(Para, v, w, target):

    '''Initialization'''
    '''---------------------------------------------------------------------'''
    # 1. Build gurobi model
    model = gp.Model()

    # 1) Variables
    y_line = model.addVars(Para.N_line, Para.N_scen, vtype = gp.GRB.BINARY)
    # 2) Operating variable
    f_flow = model.addVars(N_F_var, Para.N_scen, lb = -1e5)
    v_flow = model.addVars(N_N_var, Para.N_hour, lb = -1e5)
    l_func = model.addVars(N_L_var, Para.N_hour, lb = -1e5)
    # 3) Set objective
    h_cost = model.addVars(1, Para.N_hour, lb = -1e5)

    '''Build the model'''
    '''---------------------------------------------------------------------'''
    # 1. Constraints
    # 1) operating constraints
    for h in range(Para.N_hour):
        # Reconfiguration model
        scen = np.divmod(h, Para.N_hour / Para.N_scen)
        s = int(scen[0])
        # Select variables
        y_line_sl = y_line.select('*', s)  # y_line (reconfiguration)
        f_flow_sl = f_flow.select('*', s)  # f_flow (reconfiguration)
        v_flow_sl = v_flow.select('*', h)  # v_flow (operation)
        l_func_sl = l_func.select('*', h)  # l_func (operation)
        h_cost_sl = h_cost.select('*', h)  # h_cost (operation)
        target_sl = target[:, h]           # target (operation)
        if int(scen[1]) == 0:
            model = Reconf_Model(model, Para, y_line_sl, f_flow_sl)
        if int(scen[1]) >= 0:
            model = Operat_Model(model, Para, y_line_sl, v_flow_sl, l_func_sl, 
                                 h_cost_sl, target_sl, h, v, w)
    
    # 2) charging power constraints
    charge = gp.LinExpr()
    for n in range(Para.N_park):
        for h in range(Para.N_hour):
            charge = charge + v_flow[N_P_park + n, h]
    model.addConstr(charge >= np.sum(Para.Ecar[:, 4]) / 1000)

    # 2. Objective
    obj = gp.quicksum(h_cost)
    model.setObjective(obj, gp.GRB.MINIMIZE)

    '''Optimization'''
    '''---------------------------------------------------------------------'''
    # Set parameters
    print("\n Solving the MIP model...")
    model.optimize()
    # Optimal
    if model.status == gp.GRB.Status.OPTIMAL:
        sol = Result_MIP(model, Para, y_line, f_flow, v_flow, l_func, h_cost)
        det = Detect(Para, sol)
        # fig = Figure(Para)
        # fig.figuring(Para, sol, 0)
    else:
        sol = -1
    return sol


# This function defines the reconfiguration model
#
def Reconf_Model(model, Para, y_line, f_flow):
    
    # Constraint

    # 0. Fictitious power flow
    for n in range(Para.N_line):
        model.addConstr(f_flow[N_F_line + n] >= -1e2 * y_line[n])
        model.addConstr(f_flow[N_F_line + n] <=  1e2 * y_line[n])
    for n in range(Para.N_sub):
        model.addConstr(f_flow[N_F_sub  + n] >=  0)
        model.addConstr(f_flow[N_F_sub  + n] <=  1e2)
    for n in range(Para.N_bus):
        model.addConstr(f_flow[N_F_load + n] ==  1)
    for n in range(Para.N_gen):
        model.addConstr(f_flow[N_F_gen  + n] == -1)

    # 1. Connectivity
    for n in range(Para.N_bus):
        # Bus-branch information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - f_flow[N_F_load + n]
        expr = expr - gp.quicksum(f_flow[N_F_line + i] for i in line_head)
        expr = expr + gp.quicksum(f_flow[N_F_line + i] for i in line_tail)
        if n in Para.Sub[:,1]:
            i = int(np.where(n == Para.Sub[:,1])[0])
            expr = expr + f_flow[N_F_sub + i]
        if n in Para.Gen[:,1]:
            i = int(np.where(n == Para.Gen[:,1])[0])
            expr = expr + f_flow[N_F_gen + i]
        model.addConstr(expr == 0)
    
    # 2. Radial topology
    model.addConstr(gp.quicksum(y_line) == Para.N_bus - Para.N_sub)
    
    # Return
    return model


# This function defines the operation model
#
def Operat_Model(model, Para, y_line, v_flow, l_func, h_cost, target, h, v, w):
    
    # Objective
    # 1. Operating cost
    opr = gp.LinExpr()
    for n in range(Para.N_sub ):  # power purchasing
        opr = opr + v_flow[N_P_sub  + n] * Para.Hour[h, 4]
    for n in range(Para.N_park):  # power selling to charging park
        opr = opr - v_flow[N_P_park + n] * Para.Hour[h, 4]
    for n in range(Para.N_bus ):  # load shedding
        opr = opr + v_flow[N_P_cut  + n] * Para.Cost_pen
        opr = opr + v_flow[N_Q_cut  + n] * Para.Cost_pen
    for n in range(Para.N_gen ):  # renewables
        opr = opr + v_flow[N_S_gen  + n] * Para.Gen[n, 3]
    # 2. Augmented Lagrangian function
    alf = gp.LinExpr()
    for n in range(Para.N_park):
        alf = alf + l_func[N_L_line + n] * v[n]
        alf = alf + l_func[N_L_quad + n] * w[n] * w[n]
    
    # 3. Set objective
    model.addConstr(h_cost[0] == opr + alf)
    
    # Constraint
    # 1. Nodal active power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - gp.quicksum(v_flow[N_P_line + i] for i in line_head)
        expr = expr + gp.quicksum(v_flow[N_P_line + i] for i in line_tail)
        expr = expr + v_flow[N_P_cut + n]
        if n in Para.Sub [:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub [:, 1])[0])
            expr = expr + v_flow[N_P_sub  + i]
        if n in Para.Gen [:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen [:, 1])[0])
            expr = expr + v_flow[N_S_gen  + i] * math.cos(Para.Factor)
        if n in Para.Park[:,1]:  # charging load
            i = int(np.where(n == Para.Park[:, 1])[0])
            expr = expr - v_flow[N_P_park + i]
        model.addConstr(expr == Para.Bus[n, 1] * Para.Hour[h, 1])
    
    # 2. Nodal reactive power balance
    for n in range(Para.N_bus):
        # Bus-Line information
        line_head = Para.Line_head[n]
        line_tail = Para.Line_tail[n]
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr - gp.quicksum(v_flow[N_Q_line + i] for i in line_head)
        expr = expr + gp.quicksum(v_flow[N_Q_line + i] for i in line_tail)
        expr = expr + v_flow[N_Q_cut + n]
        if n in Para.Sub [:,1]:  # active power input from substation
            i = int(np.where(n == Para.Sub [:, 1])[0])
            expr = expr + v_flow[N_Q_sub  + i]
        if n in Para.Gen [:,1]:  # active power input from renewables
            i = int(np.where(n == Para.Gen [:, 1])[0])
            expr = expr + v_flow[N_S_gen  + i] * math.sin(Para.Factor)
        model.addConstr(expr == Para.Bus[n, 2] * Para.Hour[h, 1])
    
    # 3. Branch flow equation
    for n in range(Para.N_line):
        bus_head = int(Para.Line[n, 1])
        bus_tail = int(Para.Line[n, 2])
        # Formulate expression
        expr = gp.LinExpr()
        expr = expr + v_flow[N_V_bus + bus_head] - v_flow[N_V_bus + bus_tail]
        expr = expr - v_flow[N_P_line + n] * Para.Line[n, 4] * 2
        expr = expr - v_flow[N_Q_line + n] * Para.Line[n, 5] * 2
        model.addConstr(expr >= -Para.Big_M * (1 - y_line[n]))
        model.addConstr(expr <=  Para.Big_M * (1 - y_line[n]))
    
    # 4. Renewables generation
    for n in range(Para.N_gen):
        expr = gp.LinExpr()
        expr = expr + v_flow[N_S_gen + n]
        expr = expr + v_flow[N_C_gen + n]
        G_type = int(Para.Gen[n, 4])
        model.addConstr(expr == Para.Gen[n, 2] * Para.Hour[h, G_type + 2])

    # 5. Lower and Upper bound
    # 1) voltage amplitutde
    for n in range(Para.N_bus):
        model.addConstr(v_flow[N_V_bus  + n] >= Para.V_min)
        model.addConstr(v_flow[N_V_bus  + n] <= Para.V_max)
    # 2) line flow
    for n in range(Para.N_line):
        smax = Para.Line[n, 3]
        # active power
        model.addConstr(v_flow[N_P_line + n] >= -smax)
        model.addConstr(v_flow[N_P_line + n] <=  smax)
        model.addConstr(v_flow[N_P_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_P_line + n] <=  y_line[n] * Para.Big_M)
        # reactive power
        model.addConstr(v_flow[N_Q_line + n] >= -smax)
        model.addConstr(v_flow[N_Q_line + n] <=  smax)
        model.addConstr(v_flow[N_Q_line + n] >= -y_line[n] * Para.Big_M)
        model.addConstr(v_flow[N_Q_line + n] <=  y_line[n] * Para.Big_M)
    # 3) substation
    for n in range(Para.N_sub):
        smax = Para.Sub[n, 2]
        model.addConstr(v_flow[N_P_sub  + n] >=  0)
        model.addConstr(v_flow[N_P_sub  + n] <=  smax)
        model.addConstr(v_flow[N_Q_sub  + n] >=  0)
        model.addConstr(v_flow[N_Q_sub  + n] <=  smax)
    # 4) renewables
    for n in range(Para.N_gen):
        G_type = int(Para.Gen[n, 4])
        smax = Para.Gen[n, 2] * Para.Hour[h, G_type + 2]
        model.addConstr(v_flow[N_S_gen  + n] >=  0)
        model.addConstr(v_flow[N_S_gen  + n] <=  smax)
        model.addConstr(v_flow[N_C_gen  + n] >=  0)
        model.addConstr(v_flow[N_C_gen  + n] <=  smax)
    # 5) load shedding
    for n in range(Para.N_bus):
        smax = Para.Bus[n, 1]
        model.addConstr(v_flow[N_P_cut  + n] >=  0)
        model.addConstr(v_flow[N_P_cut  + n] <=  smax)
    for n in range(Para.N_bus):
        smax = Para.Bus[n, 2]
        model.addConstr(v_flow[N_Q_cut  + n] >=  0)
        model.addConstr(v_flow[N_Q_cut  + n] <=  smax)
    # 6) parking station
    for n in range(Para.N_park):
        model.addConstr(v_flow[N_P_park + n] >=  0)
        model.addConstr(v_flow[N_P_park + n] <=  Para.Park[n, 2])
    
    # 6. Rotation square constraints
    # 1) line flow
    for n in range(Para.N_line):
        smax = Para.Line[n, 3]
        # P_line + Q_line
        expr = v_flow[N_P_line + n] + v_flow[N_Q_line + n]
        model.addConstr(expr >= -1.414 * smax)
        model.addConstr(expr <=  1.414 * smax)
        # P_line - Q_line
        expr = v_flow[N_P_line + n] - v_flow[N_Q_line + n]
        model.addConstr(expr >= -1.414 * smax)
        model.addConstr(expr <=  1.414 * smax)
    # 2) substation
    for n in range(Para.N_sub):
        smax = Para.Sub[n, 2]
        # P_sub + Q_sub
        expr = v_flow[N_P_sub  + n] + v_flow[N_Q_sub  + n]
        model.addConstr(expr >= -1.414 * smax)
        model.addConstr(expr <=  1.414 * smax)
        # P_sub - Q_sub
        expr = v_flow[N_P_sub  + n] - v_flow[N_Q_sub  + n]
        model.addConstr(expr >= -1.414 * smax)
        model.addConstr(expr <=  1.414 * smax)
    
    # 7. Consistency Constraints
    for n in range(Para.N_park):
        difference = v_flow[N_P_park + n] - target[n]
        model.addConstr(l_func[N_L_line + n] == difference)
        model.addConstr(l_func[N_L_quad + n] >= difference * difference)

    # Return
    return model


# This function defines the COP_scheduling
#
def COP_scheduling(Para, v, w, target, flag = 0):
    
    '''Initialization'''
    '''---------------------------------------------------------------------'''
    # 1. Pre-processing
    # 1) system parameters
    capacity = ((Para.Park[: , 2] * 1000).astype(int)).tolist()
    piletype = ((Para.Park[0 : 2, 3 : 6]).astype(int)).tolist()
    # 2) Build IBM CP model
    model = cp.CpoModel()
    
    # 2. Declarations of charging variables
    # 1) interval variables of charging demand
    x_ecar = {}  # charging
    y_ecar = {}  # charging + park
    z_ecar = {}  # charging + park + rate
    for n in range(Para.N_ecar):
        time_0 = int(Para.Ecar[n, 1])
        time_1 = int(Para.Ecar[n, 2])
        x_ecar[n] = model.interval_var(
            start = [time_0, cp.INTERVAL_MAX],
            end   = [0, time_1],
            optional = False
        )
        for i in range(Para.N_park):
            y_ecar[n, i] =  model.interval_var(
                start = [time_0, cp.INTERVAL_MAX],
                end   = [0, time_1],
                optional = True
            )
            for k in range(Para.N_rate):
                z_ecar[n, i, k] = model.interval_var(
                    start = [time_0, cp.INTERVAL_MAX],
                    end   = [0, time_1],
                    size  =  int(Para.Ecar[n, 6 + k]),
                    optional = True
                )

    # 2) active power of charging park at each hour
    p_park = {}
    for i in range(Para.N_park):
        for h in range(Para.N_hour):
            p_park[i, h] = model.integer_var(0, capacity[i])

    '''Build the model'''
    '''---------------------------------------------------------------------'''
    # 1. Add constraint
    # 1) alternative constraint
    for n in range(Para.N_ecar):
        model.add(model.alternative(
            x_ecar[n], [y_ecar[n, i] for i in range(Para.N_park)]))
        for i in range(Para.N_park):
            model.add(model.alternative(
                y_ecar[n, i], [z_ecar[n, i, k] for k in range(Para.N_rate)]))
    
    # 2) capacity constraint
    charge = {}
    for i in range(Para.N_park):
        charge[i] = []
        for n in range(Para.N_ecar):
            for k in range(Para.N_rate):
                charge[i].append(model.pulse(z_ecar[n, i, k], Para.C_rate[k]))
        model.add(model.sum(charge[i]) <= capacity[i])

    # 3) pile number constraint
    occupy = {}
    for i in range(Para.N_park):
        occupy[i] = [[] for k in range(Para.N_rate)]
        for n in range(Para.N_ecar):
            for k in range(Para.N_rate):
                occupy[i][k].append(model.pulse(z_ecar[n, i, k], 1))
        for k in range(Para.N_rate):
            model.add(model.sum(occupy[i][k]) <= piletype[i][k])
    
    # 4) charging rate
    for n in range(Para.N_ecar):
        for i in range(Para.N_park):
            for k in range(Para.N_rate):
                if Para.Ecar[n, 6 + k] == 1000:
                    model.add(model.presence_of(z_ecar[n, i, k]) == 0)

    # 5) active power usage constraint
    for i in range(Para.N_park):
        for h in range(Para.N_hour):
            time_0 = int(Para.N_sect / 24 * (h + 0))  # time of start
            time_1 = int(Para.N_sect / 24 * (h + 1))  # time of end
            power = model.sum(charge[i])
            power = power + model.pulse((time_0, time_1), capacity[i])
            model.add(p_park[i, h] + capacity[i] >= power)
    
    # 2. Add objective
    obj = 0
    '''
    # 1) cost of charging park alternation
    for n in range(Para.N_ecar):
        for i in range(Para.N_park):
            if i != Para.Ecar[n, 5]:
                obj = obj + model.presence_of(y_ecar[n, i]) * Para.Cost_alt
    
    # 2) cost of timeout
    for n in range(Para.N_ecar):
        out = model.end_of(x_ecar[n]) - Para.Ecar[n, 2]
        obj = obj + model.max([out, 0]) * Para.Cost_out
    '''
    # 3) cost of power purchasing
    for i in range(Para.N_park):
        for h in range(24):
            obj = obj + p_park[i, h]
    '''
    # 4) augmented lagrange function
    if flag == 0:
        for i in range(Para.N_park):
            for h in range(Para.N_hour):
                difference = p_park[i, h] - int(target[i, h] * 1000)
                obj = obj + difference * v[i]
                obj = obj + difference * difference * w[i] * w[i] / 1e6
    '''
    obj = obj + model.max(model.end_of(x_ecar[n]) for n in range(Para.N_ecar))

    # 5) sum of all costs
    model.add(model.minimize(obj))

    '''Build the model'''
    '''---------------------------------------------------------------------'''

    # 1. Call the solver
    print("\n Solving the COP model...")
    msol = model.solve(url = None, key = None)
    print("done")

    # 2. Save the result
    status = msol.solve_status
    if status == "Optimal" or status == "Feasible":
        sol = Result_COP(msol, Para, x_ecar, y_ecar, z_ecar, p_park)
        # sol.save_scheduling(sol.scheduling)
    else:
        sol = -1
    return sol


# Entrance of the main function
#
if __name__ == "__main__":

    # Parameters
    Flag = 0
    Ncar = 50
    Name = "data/IEEE-33-parameters.xlsx"
    Para = Parameter(Name, Flag, Ncar)
    Gval = Var_index(Para)

    # Main function
    sol_ATC = ATC_algorithm(Para)