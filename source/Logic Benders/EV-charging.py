# -----------------------------------------------------------------------------
# Copyright 2020, Southeast University, Liu Pengxiang
# 
# A Hybrid MIP-CP framework for EV charging scheduling problem
# 
# EV data can be downloaded from: https://pod-point.com/guides/vehicles
# -----------------------------------------------------------------------------


import sys
import math
import time
import numpy as np
import gurobipy as gp
import openpyxl as pyxl
import docplex.cp.model  as cp
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.gridspec as gridspec
import docplex.cp.utils_visu as visu


# This class creates the System class
# -----------------------------------------------------------------------------
# 
class System(object):

    # Initialization
    def __init__(self, filename):
        # 1. pre-processing
        Tool = Excel_tool()  # excel file tool
        Data = Tool.read_excel(filename)
        # 2. data input
        self.Hour = Data[0]  # power limit at each hour
        self.Time = Data[1]  # arrival and departure distribution
        self.Park = Data[2]  # charging and parking station
        self.Pile = Data[3]  # charging pile
        self.Type = Data[4]  # type of EV battery
        self.N_hour = len(self.Hour)  # number of hours
        self.N_time = len(self.Time)  # number of times
        self.N_park = len(self.Park)  # number of parks
        self.N_pile = len(self.Pile)  # number of piles
        self.N_type = len(self.Type)  # number of types
        # 3. system parameter
        self.Cost_adjust = 20    # cost of dispatch
        self.Cost_charging = 1.26  # cost of charging
        self.Cost_delay    = 5     # cost of delay
        # 4. delay-time tolerance
        self.Time_tolerance = 2  # tolerance of time delay


# This class creates the Params class
# -----------------------------------------------------------------------------
# 
class Params(System):

    # Initialization
    def __init__(self, file_system, file_demand, flag = 0):
        # 1. inherit
        super().__init__(file_system)
        Tool = Excel_tool()  # excel file tool
        Data = Tool.read_excel(file_demand)
        # 2. parameter for demand
        self.S_rate = [50, 22, 7]  # set of charging rate
        self.N_mins = 5  # number of minutes for each time section
        self.N_sect = int(24 * 60 / self.N_mins)  # number of time section
        self.N_rate = len(self.S_rate)  # number of charging rate types
        if flag != 0:
            self.Ecar = self.EV_charging(flag)
            self.N_ecar = len(self.Ecar)
            sheet = 'Demand'
            title = ["No.", "Arrive", "Leave", "Type", "Demand", "Park", 
                "Will", "Time-50", "Time-22", "Time-07"]
            Tool.save_excel(file_demand, sheet, title, self.Ecar)
        else:
            self.Ecar = Data[0]
            self.N_ecar = len(self.Ecar)
    
    # charging demand
    def EV_charging(self, number):
        charging = np.zeros((number, 10))
        for n in range(number):
            # generate charging demand
            [time_arr, time_dep] = self.timing_generator()
            [index, demand, timing] = self.demand_generator(time_arr, time_dep)
            # write parameters
            charging[n, 0] = int(n)  # number
            charging[n, 1] = time_arr  # time of arrival
            charging[n, 2] = time_dep  # time of departure
            charging[n, 3] = index  # type of battery
            charging[n, 4] = round(demand, 3)  # charging demand
            charging[n, 5] = self.roulette(self.Park[:, 1])  # prefered park
            charging[n, 6] = np.random.randint(0, 2)  # willing of replacement
            for i, item in enumerate(timing):  # time of charging
                charging[n, 7 + i] = item
        return charging
    
    # generate the time of arrival and departure
    def timing_generator(self):
        while True:
            # time of arrival and departure
            hour_arr = self.roulette(self.Time[:, 1])
            hour_dep = self.roulette(self.Time[:, 2])
            # offset
            extr_arr = np.random.randint(0, self.N_sect / 24 + 0.1)  
            extr_dep = np.random.randint(0, self.N_sect / 24 + 0.1)
            # actual
            time_arr = int(hour_arr * self.N_sect / 24 + extr_arr)
            time_dep = int(hour_dep * self.N_sect / 24 + extr_dep)
            # charging at least 20 minutes
            if time_arr <= time_dep - 20 / self.N_mins:
                break
        return [time_arr, time_dep]
    
    # generate the charging demand of each EV
    def demand_generator(self, time_arr, time_dep):
        while True:
            # Demand
            index  = int(self.roulette(self.Type[:, 6]))
            mean   = self.Type[index, 1] / 2  # mean value
            var    = self.Type[index, 1] / 2 / 1.96  # variance
            demand = np.random.normal(loc = mean, scale = var)
            # charging time
            timing = np.zeros(self.N_rate)  # charging capacity
            for i, rate in enumerate(self.S_rate):
                flag = self.Type[index, 2 + i]
                time = demand / rate * 60 / self.N_mins
                time = min(time, time_dep - time_arr)
                if flag == 1:
                    timing[i] = max(int(time), 1)
                else:
                    timing[i] = int(-1)
            # maximum of charging demand
            list_demand = np.zeros(self.N_rate)
            for i, item in enumerate(self.S_rate):
                time = time_dep - time_arr
                flag = self.Type[index, 2 + i]
                list_demand[i] = time * flag * item / (self.N_sect / 24)
            max_demand = np.max(list_demand)
            if demand > 0 and demand < max_demand:
                break
        return [index, demand, timing]
    
    # roulette algorithm
    def roulette(self, pdf):
        pdf  = np.append(pdf, 0)
        pdf  = pdf / np.sum(pdf)  # normal scaler
        pick = np.random.rand()
        _sum = 0
        for [index, value] in enumerate(pdf):
            if pick >= _sum:
                _sum = _sum + value
            else:
                break
        solution = index - 1
        return solution


# This class creates the execution tool for Excel files
# -----------------------------------------------------------------------------
# 
class Excel_tool(object):

    # Initialization
    def __init__(self):
        pass

    # inputs data from Excel file
    def read_excel(self, filename):
        data = []
        book = pyxl.load_workbook(filename)
        # Data preprocessing
        for i, name in enumerate(book.sheetnames):  # sheet number
            if i < len(book.sheetnames):
                sheet = book[name]
                n_row = sheet.max_row     # number of rows
                n_col = sheet.max_column  # number of columns
                data.append(self.tool_filter(sheet, n_row, n_col))
        return data

    # saving data to excel file
    def save_excel(self, filename, sheetname, title, data):
        book  = pyxl.load_workbook(filename)
        name  = book.sheetnames[-1]
        book.remove(book[name])
        sheet = book.create_sheet(sheetname)
        sheet.append(title)
        for i in range(len(data)):
            sheet.append(data[i,:].tolist())  # write data
        book.save(filename = filename)
    
    # filter data from the original numpy array
    def tool_filter(self, sheet, n_row, n_col):
        k = 0
        data = []
        for i in range(n_row):
            if sheet['A' + str(i + 1)].data_type == 'n':  # if it is a number
                data.append([])
                for j in range(n_col):
                    pos = chr(64 + j + 1) + str(i + 1)  # the position
                    val = sheet[pos].value
                    if sheet[pos].data_type == 'n':
                        data[k].append(val)
                k = k + 1
        return np.array(data)


# This class creates the Result class
# -----------------------------------------------------------------------------
# 
class Result(object):
    
    # Initialization
    def __init__(self):
        pass

    # Obtain the value of variables
    def MIP_get_value(self, model, var, var_type):
        # Get value
        key = var.keys()
        val = var.copy()
        for i in range(len(key)):
            val[key[i]] = var[key[i]].x
        # Calculate dimention
        if isinstance(max(key),tuple):  # multi dimention
            dim = tuple([item + 1 for item in max(key)])
        if isinstance(max(key),int):    # one   dimention
            dim = tuple([int(len(key)),1])
        # Convert dictionary to numpy array
        arr = np.zeros(dim, dtype = var_type)
        for i in range(len(val)):
            arr[key[i]] = val[key[i]]
        return arr
    
    # Obtain the value of variables
    def COP_get_value(self, msol, var, var_type):
        # Get value
        sol = {}
        for n in list(var.keys()):
            if var_type == "int":
                sol[n] = msol.get_var_solution(var[n]).value
            if var_type == "itv":
                sol[n] = msol.get_var_solution(var[n])
        # Convert dictionary to numpy array
        return sol
    
    # Get the scheduling decisions
    def COP_get_schedule(self, sol):
        pass


# This function defines the all-in-one optimization
# -----------------------------------------------------------------------------
#
def AIO_main(Para, ev_here, ev_wait, ev_pool, sc_plan):

    # 1. Build gurobi model
    model  = gp.Model()

    # 2. Add variables
    x_ecar = model.addVars(len(ev_pool), Para.N_park, Para.N_rate, vtype = 'B')
    x_stat = model.addVars(len(ev_pool), Para.N_sect, vtype = 'B')
    x_init = model.addVars(len(ev_pool), Para.N_sect, vtype = 'B')
    f_cost = model.addVars(1)

    # 3. Pre-processing
    ev_info = [ev_here, ev_wait, ev_pool]

    # 4. Add constraints
    model = MIP_dispatch(Para, model, x_ecar, ev_info, sc_plan)

    # 5. Add objective
    model = MIP_schedule(Para, model, x_ecar, x_stat, x_init, f_cost, ev_info, sc_plan)

    # 5. Add objective
    model = MIP_minimize(Para, model, x_ecar, f_cost, ev_info)

    # 6. Solve
    # 1) call the solver
    print("\n Solving the AIO model...")
    # model.params.NonConvex = 2
    model.optimize()

    # 2) diagnostic
    if model.status == gp.GRB.Status.OPTIMAL:
        sol = Result()
        sol.x_ecar = sol.MIP_get_value(model, x_ecar, "int")
    else:
        sol = -1
    return sol, model.Runtime


# This function defines the logic based benders decomposition
# -----------------------------------------------------------------------------
#
def LBD_main(Para, ev_here, ev_wait, ev_pool, sc_plan):

    # 1. Build gurobi model
    model  = gp.Model()

    # 2. Add variables
    # 1) name
    x_ecar = model.addVars(len(ev_pool), Para.N_park, Para.N_rate, vtype = 'B')
    f_cost = model.addVars(Para.N_park, 1)

    # 3. Pre-processing
    ev_info = [ev_here, ev_wait, ev_pool]

    # 4. Add constraints
    model = MIP_dispatch(Para, model, x_ecar, ev_info, sc_plan)

    # 5. Add objective
    model = MIP_minimize(Para, model, x_ecar, f_cost,  ev_info)

    # 6. Solve
    # 1) call the solver
    print("\n Solving the LBD model...")
    model._n_call = 0
    model._x_ecar = x_ecar
    model._f_cost = f_cost
    model._x_intv = np.zeros((len(ev_pool), 3))
    model.Params.lazyConstraints = 1
    model.optimize(Lazy_Constraint_Callback)

    # 2) diagnostic
    if model.status == gp.GRB.Status.OPTIMAL:
        sol = Result()
        sol.x_intv = model._x_intv
        sol.x_ecar = sol.MIP_get_value(model, x_ecar, "int")
        sol.f_cost = sol.MIP_get_value(model, f_cost, "float")
    else:
        sol = -1
    return sol, model.Runtime


# This function optimizes the assignment of each EV
# -----------------------------------------------------------------------------
#
def MIP_dispatch(Para, model, x_ecar, ev_info, sc_plan):

    # 1. Pre-processing
    [ev_here, ev_wait, ev_pool] = ev_info

    # 2. Add constraints
    # 1) maximum type of service for each EV
    for n in range(len(ev_pool)):
        var_select = x_ecar.select(n, '*', '*')
        if n < len(ev_here):
            model.addConstr(gp.quicksum(var_select) == 1)
        else:
            model.addConstr(gp.quicksum(var_select) <= 1)

    # 2) pre-defined limits of parking assignment
    for n in range(len(ev_here)):
        index_park = sc_plan[n, 1]
        var_select = x_ecar.select(n, index_park, '*')
        model.addConstr(gp.quicksum(var_select) == 1)
    
    # 3) will of parking adjustment
    for n in range(len(ev_pool)):
        if ev_pool[n, 6] == 0:
            index_park = ev_pool[n, 5]
            var_select = x_ecar.select(n, index_park, '*')
            model.addConstr(gp.quicksum(var_select) == 1)
    
    # 4) pre-defined limits of charging rate
    for n in range(len(ev_pool)):
        for k in range(Para.N_rate):
            if ev_pool[n, k + 7] == -1:
                var_select = x_ecar.select(n, '*', k)
                model.addConstr(gp.quicksum(var_select) == 0)
    
    # 5) charging station capacity
    for p in range(Para.N_park):
        expr = gp.LinExpr()
        for n in range(len(ev_pool)):
            var_select = x_ecar.select(n, p, '*')
            expr = expr + gp.quicksum(var_select) * ev_pool[n, 4]
        model.addConstr(expr <= np.sum(Para.Hour[:, p + 1]))
    
    # 6) adjustment limit
    if len(ev_here) > 0:
        expr = gp.LinExpr()
        for n in range(len(ev_here)):
            index_park = sc_plan[n, 1]
            index_rate = sc_plan[n, 2]
            expr = expr + x_ecar[n, index_park, index_rate]
        model.addConstr(expr >= len(ev_here) * 0.995)
    
    # 7) additional constraint
    if len(ev_pool) <= np.sum(Para.Park[:, 5]):
        model.addConstr(gp.quicksum(x_ecar) == len(ev_pool))
    else:
        model.addConstr(gp.quicksum(x_ecar) >= np.sum(Para.Park[:, 5]))

    return model


# This function adds the objective function
# -----------------------------------------------------------------------------
#
def MIP_minimize(Para, model, x_ecar, f_cost, ev_info):
    
    # 1. Pre-processing
    [ev_here, ev_wait, ev_pool] = ev_info

    # 2. Add objective
    # 1) charging income
    obj_reject = gp.LinExpr()
    for n in range(len(ev_pool)):
        var_select = x_ecar.select(n, '*', '*')
        var_coef   = ev_pool[n, 4] * Para.Cost_charging
        obj_reject = obj_reject + var_coef * (1 - gp.quicksum(var_select))

    # 2) charging park adjustment cost
    obj_adjust = gp.LinExpr()
    for n in range(len(ev_pool)):
        for p in range(Para.N_park):
            if n >= len(ev_here) and ev_pool[n, 5] != p:
                var_select = x_ecar.select(n, p, '*')
                var_coef   = Para.Cost_adjust
                obj_adjust = obj_adjust + var_coef * gp.quicksum(var_select)
    
    # 3. Set objective
    objective = obj_reject + obj_adjust + gp.quicksum(f_cost)
    model.setObjective(objective, gp.GRB.MINIMIZE)

    return model


# This function optimizes the assignment of each EV
# -----------------------------------------------------------------------------
#
def MIP_schedule(Para, model, x_ecar, x_stat, x_init, f_cost, ev_info, sc_plan):
    
    # 1. Pre-processing
    [ev_here, ev_wait, ev_pool] = ev_info

    # 2. Add constraints
    # 1) charging time for each ev
    for n in range(len(ev_pool)):
        for t in range(Para.N_sect):
            if t <  ev_pool[n, 1]:
                model.addConstr(x_stat[n, t] == 0)
            if t >= ev_pool[n, 2] + Para.Time_tolerance:
                model.addConstr(x_stat[n, t] == 0)
    
    # 2) start charging
    for n in range(len(ev_pool)):
        for t in range(Para.N_sect - 1):
            model.addConstr(x_init[n, t] >= x_stat[n, t + 1] - x_stat[n, t])
    
    # 3) power demand for each ev
    for n in range(len(ev_pool)):
        expr = gp.LinExpr()
        for k in range(Para.N_rate):
            x_ecar_sl = x_ecar.select(n, '*', k)
            expr = expr + ev_pool[n, k + 7] * gp.quicksum(x_ecar_sl)
        model.addConstr(gp.quicksum(x_stat.select(n, '*')) == expr)

    # 4) no charging gap
    for n in range(len(ev_pool)):
        x_init_sl = x_init.select(n, '*')
        x_ecar_sl = x_ecar.select(n, '*', '*')
        model.addConstr(gp.quicksum(x_init_sl) <= gp.quicksum(x_ecar_sl))

    # 5) pile number
    for t in range(Para.N_sect):
        for p in range(Para.N_park):
            for k in range(Para.N_rate):
                expr = gp.QuadExpr()
                for n in range(len(ev_pool)):
                    bilinear_var = x_ecar[n, p, k] * x_stat[n, t]
                    expr = expr + bilinear_var
                model.addConstr(expr <= Para.Park[p, k + 2])

    # 6) power limit
    for t in range(Para.N_sect):
        h = int(t / (60 / Para.N_mins))
        for p in range(Para.N_park):
            expr = gp.QuadExpr()
            for k in range(Para.N_rate):
                for n in range(len(ev_pool)):
                    bilinear_var = x_ecar[n, p, k] * x_stat[n, t]
                    expr = expr + bilinear_var * Para.S_rate[k]
            model.addConstr(expr <= Para.Hour[h, p + 1])
    
    # 7) delay cost
    expr = gp.LinExpr()
    for n in range(len(ev_pool)):
        for t in range(Para.N_sect):
            if t >= ev_pool[n, 2]:
                expr = expr + x_stat[n, t] * Para.Cost_delay
    model.addConstr(gp.quicksum(f_cost) >= expr)
    
    return model


# This function optimizes the scheduling of EV charging
# -----------------------------------------------------------------------------
# 
def COP_schedule(Para, ev_pool, index_rate, p):
    
    # 1. Pre-processing
    capacity = Para.Park[p, 1]
    uppbound = Para.Hour[:, p + 1]

    # 2. Build IBM CP model
    model = cp.CpoModel()

    # 3. Add variables
    # 1) interval variables of charging demand
    x_intv = {}
    for n in range(len(ev_pool)):
        time_0 = int(ev_pool[n, 1])
        time_1 = int(ev_pool[n, 2])
        x_intv[n] = model.interval_var(
            start = [time_0, cp.INTERVAL_MAX], 
            end   = [time_0, time_1 + Para.Time_tolerance],
            size  = int(ev_pool[n, 7 + int(index_rate[n])])
        )
    
    # 4. Add constraints
    # 1) capacity constraint
    charge = []
    for n in range(len(ev_pool)):
        k = int(index_rate[n])
        charge.append(model.pulse(x_intv[n], Para.S_rate[k]))
    model.add(model.sum(charge) <= capacity)
    
    # 2) pile number constraint
    occupy = [[] for k in range(Para.N_rate)]
    for n in range(len(ev_pool)):
        k = int(index_rate[n])
        occupy[k].append(model.pulse(x_intv[n], 1))
    for k in range(Para.N_rate):
        model.add(model.sum(occupy[k]) <= Para.Park[p, k + 2])

    # 3) active power usage
    p_park = []
    for h in range(Para.N_hour):
        time_0 = int(Para.N_sect / 24 * (h + 0))  # time of start
        time_1 = int(Para.N_sect / 24 * (h + 1))  # time of end
        height = np.max(uppbound) - uppbound[h]   # height of pulse function
        if height > 0:
            p_park.append(model.pulse((time_0, time_1), height))
    model.add(model.sum(p_park) + model.sum(charge) <= np.max(uppbound))

    # 5. Add objective
    obj = 0
    for n in range(len(ev_pool)):
        obj = obj + model.max(model.end_of(x_intv[n]) - ev_pool[n, 2], 0)
    model.add(model.minimize(obj * Para.Cost_delay))

    # 6. Solve the model
    # 1) Call the solver
    msol = model.solve(
        RelativeOptimalityTolerance = 0.05,
        LogVerbosity = 'Quiet',
        FailLimit = 10000
    )

    # 2) Save the result
    status = msol.solve_status
    if status == "Optimal" or status == "Feasible":
        sol = Result()
        sol.x_intv = sol.COP_get_value(msol, x_intv, 'itv')
        sol.object = msol.get_objective_values()[0]
    else:
        sol = -1
    return sol


# This function defines the lazy constraint callback approach
# -----------------------------------------------------------------------------
# 
def Lazy_Constraint_Callback(model, where):

    # Incumbent solution
    if where == gp.GRB.Callback.MIPSOL:
        model._n_call = model._n_call + 1
        # get temperate solution
        incumb = model.cbGetSolution(model._x_ecar)
        x_ecar = np.zeros((len(ev_pool), Para.N_park, Para.N_rate))
        for n in range(len(ev_pool)):
            for p in range(Para.N_park):
                for k in range(Para.N_rate):
                    x_ecar[n, p, k] = int(round(incumb[n, p, k]))
                    
        # index
        index = np.where(x_ecar == 1)
        index_ecar = index[0]
        index_park = index[1]
        index_rate = index[2]

        # run the COP scheduling model
        error = Para.N_park
        for p in range(Para.N_park):
            id_0 = np.where(index_park == p)[0]
            id_1 = index_ecar[id_0]
            index_ecar[id_0]
            if len(id_0) > 0:
                sol = COP_schedule(Para, ev_pool[id_1, :], index_rate[id_0], p)
                expr = gp.LinExpr()
                for n in id_0:
                    k = int(index_rate[n])
                    expr = expr + 1 - model._x_ecar[n, p, k]
                # lazy constraint for arrangement
                if sol == -1:
                    model.cbLazy(expr >= 1)
                else:
                    # model.cbLazy(model._f_cost[p, 0] >= sol.object * expr)
                    error = error - 1
                    for i, item in enumerate(id_1):
                        model._x_intv[item, 0] = sol.x_intv[i].start
                        model._x_intv[item, 1] = sol.x_intv[i].end
                        model._x_intv[item, 2] = sol.x_intv[i].size

        # lazy constraint for ev number
        if error == 0:
            ev_number = len(index_park)
            if ev_number < len(ev_pool):
                model.cbLazy(gp.quicksum(model._x_ecar) >= ev_number + 1)


# This function defines the figuring
# -----------------------------------------------------------------------------
# 
def Figure(Para, ev_here, ev_wait, ev_pool, sol):
    
    # Pre-process
    fig = plt.figure(tight_layout=True)
    gs = gridspec.GridSpec(2, 2)
    
    bdata = []  # EV data in the box
    bindx = []  # index of EV
    for i in range(Para.N_park):
        bdata.append([])
        bindx.append([])
        for k in range(Para.N_rate):
            dtemp = []
            itemp = []
            index = np.where(sol.x_ecar[:,i,k] == 1)[0]
            index = index[np.where(index >= len(ev_here))[0]]
            for item in index:
                data = np.zeros(4)
                data[0] = ev_pool[item, 1]
                data[1] = ev_pool[item, 2]
                data[2] = sol.x_intv[item, 0]
                data[3] = sol.x_intv[item, 1]
                dtemp.append(data)
                itemp.append(item - len(ev_here) + 1)
            bdata[i].append(np.array(dtemp))
            bindx[i].append(np.array(itemp))

    for i in range(Para.N_park):
        
        font = {'family': 'Times New Roman'}
        if i == 0:
            ax = fig.add_subplot(gs[0, :])
        else:
            ax = fig.add_subplot(gs[1, i - 1])

        data = np.vstack(
            [bdata[i][k] for k in range(Para.N_rate) if bdata[i][k].size > 0]
        )
        indx = np.hstack(
            [bindx[i][k] for k in range(Para.N_rate) if bindx[i][k].size > 0]
        )

        box = ax.boxplot(data.T, whis = 100, patch_artist = True)
        for median in box['medians']:
            median.set(linewidth = 0)
        
        color = ['tab:blue', 'tab:orange', 'tab:green']
        power = ['50 kW', '22 kW', '7 kW']
        N_num = []
        for k in range(Para.N_rate):
            for n in range(len(bindx[i][k])):
                N_num.append(k)

        for n, patchs in enumerate(box['boxes']):
            patchs.set_color(color[N_num[n]])
            patchs.set_facecolor(color[N_num[n]])

        # axis
        ax.set_xticks(np.arange(1, len(data) + 1, 1))
        ax.set_yticks([72, 144, 216, 288])
        ax.set_xticklabels((str(i) for i in indx), font)
        ax.set_yticklabels(['6','12','18','24'], font)
        ax.set_xlabel('Index of EV assigned to station-%d' % (i+1), font)
        ax.set_ylabel('Time (h)', font)
        ax.set_ylim(50, 300)
        
        # legend
        if i == 0:
            patch = []
            for k in range(Para.N_rate):
                patch.append(mpatches.Patch(color = color[k]))
            plt.legend(
                handles = patch, labels = power, loc = 9,
                ncol = 3, prop = font,
                # bbox_to_anchor = (0.5, 1.25),
            )
    plt.show()


# Main function
# -----------------------------------------------------------------------------
# 
if __name__ == "__main__":

    # Parameters
    file_system = "data/IEEE-Simplified/Data-charging-station.xlsx"
    file_demand = "data/IEEE-Simplified/Data-EV-demand.xlsx"
    Para = Params(file_system, file_demand, 0)

    # 
    arrival = np.array([0, 500, 100, 100, 100, 50, 50, 25, 25])
    # arrival = np.array([0, 460, 50])
    arrival = np.cumsum(arrival)
    #

    sc_plan = []
    ev_here = []
    time_MP = np.zeros((len(arrival) - 1, 2))
    time_CP = np.zeros((len(arrival) - 1, 2))
    for n in range(len(arrival) - 1):
        # data preparation
        ev_wait = Para.Ecar[arrival[n] : arrival[n + 1], :]
        ev_pool = np.reshape(np.append(ev_here, ev_wait), (-1, 10))
        
        # AIO model
        t0 = time.perf_counter()
        sol, runtime = AIO_main(Para, ev_here, ev_wait, ev_pool, sc_plan)
        t1 = time.perf_counter()
        time_MP[n, 0] = t1 - t0
        time_MP[n, 1] = runtime
        
        # LBD model
        t2 = time.perf_counter()
        sol, runtime = LBD_main(Para, ev_here, ev_wait, ev_pool, sc_plan)
        t3 = time.perf_counter()
        time_CP[n, 0] = t3 - t2
        time_CP[n, 1] = runtime

        # update
        index = np.where(sol.x_ecar == 1)
        index_ecar = index[0]
        index_park = index[1]
        index_rate = index[2]
        sc_plan = np.array([index_ecar, index_park, index_rate]).T
        ev_here = ev_pool[index_ecar, :]

        # figure
        # Figure(Para, ev_here, ev_wait, ev_pool, sol)
    
    tool = Excel_tool()
    
    name = "result/time_MP.xlsx"
    tool.save_excel(name, "result", ["total", "run"], time_MP)
    
    name = "result/time_CP.xlsx"
    tool.save_excel(name, "result", ["total", "run"], time_CP)

